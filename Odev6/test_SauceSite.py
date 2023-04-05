from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl



#Bekleme işlemlerini ele alacak kütüphanedir.
from selenium.webdriver.support.wait import WebDriverWait
#İlgili bekleme işlemlerini yaparken hangi şarta olarak bekleneceğini söyler.
from selenium.webdriver.support import expected_conditions as ec



class Test_FinalProject:
    Url = "https://www.saucedemo.com/"
    usernameID = "user-name"
    passwordID = "password"
    loginBtnID = "login-button"
    productSortCss = "*[data-test='product_sort_container']"
    addToChart_BackPack = "*[data-test='add-to-cart-sauce-labs-backpack']"
    addToChat_boltTshirt = "*[data-test='add-to-cart-sauce-labs-bolt-t-shirt']" 
    cartBadge = ".shopping_cart_badge"
    checkoutID = "checkout"
    firstNameID = "first-name"
    lastNameID = "last-name"
    postalCodeID = "postal-code"
    continueID = "continue"
    finishID = "finish" 

    firstNameInput = "Mahmut Can"
    lastNameInput = "Sarıbal"
    postalCodeInput = "34000"

    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(self.Url)

        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True,)

    
    def teardown_method(self):
        self.driver.quit()
    
    def test_demo(self):
        pass
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located(locator))
    
    def getData():
        #veriyi Al
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        #Çalışacağımız sayfası söylüyoruz.
        selectedSheet = excelFile["Sheet1"]
        #Toplam satırı buluyoruz.
        totalRows = selectedSheet.max_row
        data = []

        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        
        return data
    
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        errorMsg = "Epic sadface: Username and password do not match any user in this service"
        self.waitForElementVisible((By.ID,self.usernameID))
        usernameInput = self.driver.find_element(By.ID,self.usernameID)
        self.waitForElementVisible((By.ID,self.passwordID))
        passwordInput = self.driver.find_element(By.ID,self.passwordID)

        usernameInput.send_keys(username)
        passwordInput.send_keys(password)

        loginBtn = self.driver.find_element(By.ID,self.loginBtnID).click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")

        assert errorMessage.text == errorMsg
    @pytest.mark.parametrize("username,password",[("standard_user","secret_sauce")])
    def test_CompleteShop(self,username,password):
        self.waitForElementVisible((By.ID,self.usernameID))
        usernameInput = self.driver.find_element(By.ID,self.usernameID)
        self.waitForElementVisible((By.ID,self.passwordID))
        passwordInput = self.driver.find_element(By.ID,self.passwordID)

        usernameInput.send_keys(username)
        passwordInput.send_keys(password)

        self.waitForElementVisible((By.ID,self.loginBtnID))
        loginBtn = self.driver.find_element(By.ID,self.loginBtnID).click()

        self.waitForElementVisible((By.CSS_SELECTOR,self.productSortCss))
        productSortBtn = self.driver.find_element(By.CSS_SELECTOR,self.productSortCss).click()

        self.waitForElementVisible((By.CSS_SELECTOR,self.addToChat_boltTshirt))
        addChartBolt = self.driver.find_element(By.CSS_SELECTOR,self.addToChat_boltTshirt).click()

        self.waitForElementVisible((By.CSS_SELECTOR,self.addToChart_BackPack))
        addChartBackPack = self.driver.find_element(By.CSS_SELECTOR,self.addToChart_BackPack).click()

        self.waitForElementVisible((By.CSS_SELECTOR,self.cartBadge))
        cartBadgeBtn = self.driver.find_element(By.CSS_SELECTOR,self.cartBadge).click()
        self.driver.save_screenshot(f"{self.folderPath}/Badge.png")

        self.waitForElementVisible((By.ID,self.checkoutID))
        checkBtn = self.driver.find_element(By.ID,self.checkoutID).click()

        self.waitForElementVisible((By.ID,self.firstNameID))
        firstName = self.driver.find_element(By.ID,self.firstNameID)
        firstName.send_keys(self.firstNameInput)

        self.waitForElementVisible((By.ID,self.lastNameID))
        lastName = self.driver.find_element(By.ID,self.lastNameID)
        lastName.send_keys(self.lastNameInput)

        self.waitForElementVisible((By.ID,self.postalCodeID))
        postalCodeInput = self.driver.find_element(By.ID,self.postalCodeID)
        postalCodeInput.send_keys(self.postalCodeInput)
        self.driver.save_screenshot(f"{self.folderPath}/firstname-lastname-postalcode.png")
        self.waitForElementVisible((By.ID,self.continueID))
        continueBtn = self.driver.find_element(By.ID,self.continueID).click()

        self.waitForElementVisible((By.ID,self.finishID))
        finishBtn = self.driver.find_element(By.ID,self.finishID).click()
        self.driver.save_screenshot(f"{self.folderPath}/finist.png")



