from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl

#Bekleme işlemlerini ele alacak kütüphanedir.
from selenium.webdriver.support.wait import WebDriverWait
#İlgili bekleme işlemlerini yaparken hangi şarta olarak bekleneceğini söyler.
from selenium.webdriver.support import expected_conditions as ec



class Test_DemoClass:
    #Her testen önce çağrılır.
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")

        # günün tarihi al bu tarih ile bir klasör var mı kontrol et yoksa oluştur.
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True,)
        
    #her testten sonra çağrılır.
    def teardown_method(self):
        self.driver.quit()

    def readData(self):
        print("x")
    
    # setup -> test_demoFunc -> teardown
    def test_demoFunc(self):
        # 3A Act Arrange Assert
        text = "Hello"
        assert text == "Hello"
     #setup -> test_demo2 -> teardown
    def test_demo2(self):
        assert True
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located(locator))
    
    
    def getData():
        #veriyi Al
        excelFile = openpyxl.load_workbook("/data/invalid_login.xlsx")
        #Çalışacağımız sayfası söylüyoruz.
        selectedSheet = excelFile["Sheet1"]
        #Toplam satırı buluyoruz.
        totalRows = selectedSheet.max_row
        data = []

        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupledata)
        
        return data



    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        #Maksimum 5 saniye.
                #Until; ne zamana kadar ...  şuna kadar bekle.
                #Visibility: görünür olma durumu.
        # En fazla 5 saniye olacak şekilde user-name id'li elementin görünmesini bekle.
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)

        loginBtn = self.driver.find_element(By.ID,"login-button")

        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")

        #Ekran görüntüsü alır.
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")

        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
        
    
    
